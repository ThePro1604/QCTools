import openpyxl
import PySimpleGUI as sg


def ToolRun_ExcelCells2Files():
    def _onKeyRelease(event):
        ctrl = (event.state & 0x4) != 0
        if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
            event.widget.event_generate("<<Cut>>")

        if event.keycode == 86 and ctrl and event.keysym.lower() != "v":
            event.widget.event_generate("<<Paste>>")

        if event.keycode == 67 and ctrl and event.keysym.lower() != "c":
            event.widget.event_generate("<<Copy>>")

        if event.keycode == 65 and ctrl and event.keysym.lower() != "a":
            event.widget.event_generate("<<SelectAll>>")

    def ExcelCells2Files(folder, file, document_name_cell, cell):
        wb = openpyxl.load_workbook(file)

        sheet_obj = wb.active
        cell_obj = sheet_obj.cell(row=1, column=1)

        ColNames = {}
        Current = 0
        for COL in sheet_obj.iter_cols(1, sheet_obj.max_column):
            ColNames[COL[0].value] = Current
            Current += 1
        for row in sheet_obj.rows:

            # requestID = row[ColNames['DocumentId'].value].value

            requestID = row[ColNames[document_name_cell]].value
            result = row[ColNames[cell]].value
            try:
                filename = requestID + ".txt"
            except:
                continue

            print(result)
            try:
                text_file = open(folder + "\\" + filename, "w")
                text_file.write(result)
                text_file.close()
            except:
                continue



        layout_popup = [
            [sg.Text("Done!")],
            [sg.Button('Close')]
        ]
        popup = sg.Window("Complete", layout_popup, icon=r'N:\Images\Shahaf\Projects\Assests\excel2file.ico')

        while True:
            event, values = popup.read()
            if event == "Close" or event == sg.WIN_CLOSED:
                break
        popup.close()


    left = [
        [sg.Text("Choose File")],
        [sg.FileBrowse(key="file_name",file_types=(("ALL Files", "*.*"), ("ALL CSV Files", "*.csv"), ("ALL XLSX Files", "*.xlsx"),)),
         sg.InputText(key='myfile', )],
    ]

    right = [
        [sg.Text("Output Folder")],
        [sg.FolderBrowse(key="folder_name"), sg.InputText(key='myfolder')],

    ]

    middle = [
        [sg.Text("DocumentID cell"), sg.InputText(key='document_id', size=(10, 1))],
        [sg.Text("Exported Cell"), sg.InputText(key='cell', size=(13, 1))],
    ]

    buttons = [
        [sg.Button('Start'), sg.Button('Close')],
    ]

    copyright = [
        [sg.Text("Ⓒ By Shahaf Stossel", text_color="#A20909")],
    ]

    description = [
        [sg.Text("This tool is used to create separate files out of excel cells in a specific column")],
        [sg.Text("the DocumentID cell is the one that is being used as the Name of the File")]

    ]


    layout = [
        [sg.Push(), sg.Column(description), sg.Push()],
        [sg.HSeparator()],
        [
            sg.Column(left, vertical_alignment="top", key='-COL1-'),
            sg.VSeparator(),
            sg.Column(right, vertical_alignment="top", key='-COL1-')
        ],
        [sg.VPush()],
        [sg.Push(), sg.Column(middle), sg.Push()],
        [sg.VPush()],
        [sg.Push(), sg.Column(buttons), sg.Push()],
        [sg.Push(), sg.Column(copyright), sg.Push()],

    ]

    window = sg.Window("ExcelCells2Files", layout, icon=r'N:\Images\Shahaf\Projects\Assests\excel2file.ico', finalize=True)
    window.TKroot.bind_all("<Key>", _onKeyRelease, "+")

    while True:
        event, values = window.read()
        if event == "Start" and len(values['myfolder']) > 1:
            ExcelCells2Files(values['myfolder'], values['file_name'], values['document_id'], values['cell'])
        if event == sg.WIN_CLOSED or event == "Close":
            break

    window.close()


ToolRun_ExcelCells2Files()
